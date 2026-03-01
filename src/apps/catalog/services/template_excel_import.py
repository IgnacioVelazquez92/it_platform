from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import BinaryIO
import unicodedata

from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from openpyxl import load_workbook

from apps.catalog.forms.helpers_globals import save_globals_for_selection_set
from apps.catalog.models import (
    AccessTemplate,
    AccessTemplateItem,
    ActionPermission,
    Company,
    ErpModule,
    ErpModuleLevel,
    ErpModuleSubLevel,
    PermissionSelectionSet,
)
from apps.catalog.models.permissions.global_ops import ActionValueType
from apps.catalog.models.selections import SelectionSetLevel, SelectionSetModule, SelectionSetSubLevel

MODULES_START_ROW = 2
MODULES_END_ROW = 449
ACTIONS_START_ROW = 451
ACTIONS_END_ROW = 542

YES_VALUES = {"si", "s", "yes", "y", "true", "1", "x"}
NO_VALUES = {"no", "n", "false", "0", ""}


class TemplateExcelImportError(Exception):
    pass


@dataclass(frozen=True)
class SheetImportResult:
    sheet_name: str
    template_id: int
    created: bool
    modules_selected: int
    sublevels_selected: int
    action_values_selected: int


@dataclass(frozen=True)
class ParsedSheetTemplate:
    name: str
    modules: list[ErpModule]
    levels: list[ErpModuleLevel]
    sublevels: list[ErpModuleSubLevel]
    action_items: list[dict]
    action_values_selected: int


@dataclass(frozen=True)
class WorkbookImportResult:
    results: list[SheetImportResult]
    warnings: list[str] = field(default_factory=list)


def _normalize_text(value) -> str:
    text = " ".join(str(value or "").strip().split())
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.casefold()


def _read_file_bytes(file_obj: UploadedFile | BinaryIO) -> bytes:
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    content = file_obj.read()
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    return content


def _as_bool(value, *, sheet_name: str, row_idx: int, label: str) -> bool:
    norm = _normalize_text(value)
    if norm in YES_VALUES:
        return True
    if norm in NO_VALUES:
        return False
    raise TemplateExcelImportError(
        f"Solapa '{sheet_name}', fila {row_idx}: valor invalido para '{label}': {value!r}."
    )


def _as_int(value, *, sheet_name: str, row_idx: int, label: str) -> int | None:
    if value in (None, ""):
        return None
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise TemplateExcelImportError(
            f"Solapa '{sheet_name}', fila {row_idx}: valor invalido para '{label}': {value!r}."
        )
    if decimal_value != decimal_value.to_integral_value():
        raise TemplateExcelImportError(
            f"Solapa '{sheet_name}', fila {row_idx}: se esperaba un entero para '{label}' y llego {value!r}."
        )
    return int(decimal_value)


def _as_decimal(value, *, sheet_name: str, row_idx: int, label: str) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        raise TemplateExcelImportError(
            f"Solapa '{sheet_name}', fila {row_idx}: valor invalido para '{label}': {value!r}."
        )


def _build_module_index() -> dict[tuple[str, str, str], ErpModuleSubLevel]:
    sublevels = (
        ErpModuleSubLevel.objects.filter(
            is_active=True,
            level__is_active=True,
            level__module__is_active=True,
        )
        .select_related("level", "level__module")
        .order_by("level__module__name", "level__name", "name")
    )
    index: dict[tuple[str, str, str], ErpModuleSubLevel] = {}
    duplicates: set[tuple[str, str, str]] = set()
    for sublevel in sublevels:
        key = (
            _normalize_text(sublevel.level.module.name),
            _normalize_text(sublevel.level.name),
            _normalize_text(sublevel.name),
        )
        if key in index:
            duplicates.add(key)
            continue
        index[key] = sublevel
    if duplicates:
        raise TemplateExcelImportError(
            "Hay subniveles ERP duplicados luego de normalizar nombres. Revisar catalogo base."
        )
    return index


def _build_level_index() -> dict[tuple[str, str], ErpModuleLevel]:
    levels = (
        ErpModuleLevel.objects.filter(
            is_active=True,
            module__is_active=True,
        )
        .select_related("module")
        .order_by("module__name", "name")
    )
    index: dict[tuple[str, str], ErpModuleLevel] = {}
    duplicates: set[tuple[str, str]] = set()
    for level in levels:
        key = (
            _normalize_text(level.module.name),
            _normalize_text(level.name),
        )
        if key in index:
            duplicates.add(key)
            continue
        index[key] = level
    if duplicates:
        raise TemplateExcelImportError(
            "Hay niveles ERP duplicados luego de normalizar nombres. Revisar catalogo base."
        )
    return index


def _build_root_module_index() -> dict[str, ErpModule]:
    modules = ErpModule.objects.filter(is_active=True).order_by("name")
    index: dict[str, ErpModule] = {}
    duplicates: set[str] = set()
    for module in modules:
        key = _normalize_text(module.name)
        if key in index:
            duplicates.add(key)
            continue
        index[key] = module
    if duplicates:
        raise TemplateExcelImportError(
            "Hay modulos ERP duplicados luego de normalizar nombres. Revisar catalogo base."
        )
    return index


def _build_action_index() -> dict[tuple[str, str], ActionPermission]:
    actions = ActionPermission.objects.filter(is_active=True).order_by("group", "action")
    index: dict[tuple[str, str], ActionPermission] = {}
    duplicates: set[tuple[str, str]] = set()
    for action in actions:
        key = (_normalize_text(action.group), _normalize_text(action.action))
        if key in index:
            duplicates.add(key)
            continue
        index[key] = action
    if duplicates:
        raise TemplateExcelImportError(
            "Hay ActionPermission duplicados luego de normalizar nombres. Revisar catalogo base."
        )
    return index


def _parse_sheet(sheet, *, module_index, level_index, root_module_index, action_index) -> ParsedSheetTemplate:
    selected_modules: dict[int, ErpModule] = {}
    selected_levels: dict[int, ErpModuleLevel] = {}
    selected_sublevels: dict[int, ErpModuleSubLevel] = {}
    action_items: list[dict] = []
    action_values_selected = 0

    for row_idx in range(MODULES_START_ROW, MODULES_END_ROW + 1):
        module_name = sheet.cell(row=row_idx, column=1).value
        level_name = sheet.cell(row=row_idx, column=2).value
        sublevel_name = sheet.cell(row=row_idx, column=3).value
        applies_value = sheet.cell(row=row_idx, column=4).value

        if not any([module_name, level_name, sublevel_name, applies_value]):
            continue

        if not _as_bool(applies_value, sheet_name=sheet.title, row_idx=row_idx, label="Corresponde"):
            continue

        module_key = _normalize_text(module_name)
        level_key = _normalize_text(level_name)
        sublevel_key = _normalize_text(sublevel_name)

        module = root_module_index.get(module_key)
        if not module:
            raise TemplateExcelImportError(
                f"Solapa '{sheet.title}', fila {row_idx}: no existe el modulo {module_name!r}."
            )
        selected_modules[module.id] = module

        if not level_key:
            continue

        level = level_index.get((module_key, level_key))
        if not level:
            raise TemplateExcelImportError(
                f"Solapa '{sheet.title}', fila {row_idx}: no existe el nivel "
                f"{module_name!r} / {level_name!r}."
            )
        selected_levels[level.id] = level

        if not sublevel_key:
            continue

        sublevel = module_index.get((module_key, level_key, sublevel_key))
        if not sublevel:
            raise TemplateExcelImportError(
                f"Solapa '{sheet.title}', fila {row_idx}: no existe el subnivel "
                f"{module_name!r} / {level_name!r} / {sublevel_name!r}."
            )
        selected_sublevels[sublevel.id] = sublevel

    for row_idx in range(ACTIONS_START_ROW, ACTIONS_END_ROW + 1):
        group_name = sheet.cell(row=row_idx, column=1).value
        action_name = sheet.cell(row=row_idx, column=2).value
        raw_value = sheet.cell(row=row_idx, column=3).value

        if not any([group_name, action_name, raw_value]):
            continue

        key = (_normalize_text(group_name), _normalize_text(action_name))
        action_permission = action_index.get(key)
        if not action_permission:
            raise TemplateExcelImportError(
                f"Solapa '{sheet.title}', fila {row_idx}: no existe el permiso global "
                f"{group_name!r} / {action_name!r}."
            )

        item = {"action_permission": action_permission}
        value_type = action_permission.value_type

        if value_type == ActionValueType.BOOL:
            value_bool = _as_bool(raw_value, sheet_name=sheet.title, row_idx=row_idx, label=action_name)
            item["value_bool"] = value_bool
            if value_bool:
                action_values_selected += 1
        elif value_type == ActionValueType.INT:
            value_int = _as_int(raw_value, sheet_name=sheet.title, row_idx=row_idx, label=action_name)
            item["value_int"] = value_int
            if value_int is not None:
                action_values_selected += 1
        elif value_type in (ActionValueType.DECIMAL, ActionValueType.PERCENT):
            value_decimal = _as_decimal(raw_value, sheet_name=sheet.title, row_idx=row_idx, label=action_name)
            item["value_decimal"] = value_decimal
            if value_decimal is not None:
                action_values_selected += 1
        else:
            value_text = " ".join(str(raw_value or "").strip().split())
            item["value_text"] = value_text
            if value_text:
                action_values_selected += 1

        action_items.append(item)

    return ParsedSheetTemplate(
        name=" ".join(sheet.title.strip().split()),
        modules=sorted(selected_modules.values(), key=lambda row: row.name),
        levels=sorted(selected_levels.values(), key=lambda row: (row.module.name, row.name)),
        sublevels=sorted(
            selected_sublevels.values(),
            key=lambda row: (row.level.module.name, row.level.name, row.name),
        ),
        action_items=action_items,
        action_values_selected=action_values_selected,
    )


def _delete_template_items_and_orphans(template: AccessTemplate) -> None:
    selection_set_ids = list(template.items.values_list("selection_set_id", flat=True))
    if template.selection_set_id:
        selection_set_ids.append(template.selection_set_id)
        template.selection_set = None
        template.save(update_fields=["selection_set"])
    template.items.all().delete()
    PermissionSelectionSet.objects.filter(
        pk__in=selection_set_ids,
        template_items__isnull=True,
        request_items__isnull=True,
        templates_legacy__isnull=True,
    ).delete()


def _resolve_company(*, company: Company | None) -> Company:
    if company is not None:
        return company

    resolved = Company.objects.filter(is_active=True).order_by("name").first()
    if resolved is None:
        raise TemplateExcelImportError("No hay empresas activas para crear el item base del template.")
    return resolved


@transaction.atomic
def import_templates_from_excel(
    *,
    file_obj: UploadedFile | BinaryIO,
    owner,
    company: Company | None = None,
    replace_existing: bool = False,
) -> WorkbookImportResult:
    if owner is None or getattr(owner, "pk", None) is None:
        raise TemplateExcelImportError("Se requiere un usuario owner valido para la importacion.")

    base_company = _resolve_company(company=company)
    workbook = load_workbook(filename=BytesIO(_read_file_bytes(file_obj)), data_only=True)
    module_index = _build_module_index()
    level_index = _build_level_index()
    root_module_index = _build_root_module_index()
    action_index = _build_action_index()

    parsed_sheets: list[ParsedSheetTemplate] = []
    warnings: list[str] = []

    for sheet in workbook.worksheets:
        template_name = " ".join(sheet.title.strip().split())
        if not template_name:
            warnings.append("Se encontro una solapa sin nombre y fue omitida.")
            continue
        parsed = _parse_sheet(
            sheet,
            module_index=module_index,
            level_index=level_index,
            root_module_index=root_module_index,
            action_index=action_index,
        )
        parsed_sheets.append(parsed)

    results: list[SheetImportResult] = []
    for parsed in parsed_sheets:
        template = AccessTemplate.objects.filter(name=parsed.name).first()
        created = template is None

        if template and not replace_existing:
            raise TemplateExcelImportError(
                f"Ya existe un template llamado '{parsed.name}'. Activar reemplazo para actualizarlo."
            )

        if template is None:
            template = AccessTemplate.objects.create(
                name=parsed.name,
                department="",
                role_name="",
                notes="Importado desde Excel.",
                owner=owner,
                is_active=True,
            )
        else:
            _delete_template_items_and_orphans(template)
            template.owner = owner
            template.is_active = True
            if not template.notes:
                template.notes = "Importado desde Excel."
            template.save(update_fields=["owner", "is_active", "notes"])

        selection_set = PermissionSelectionSet.objects.create(
            company=base_company,
            branch=None,
            notes=f"Importado desde Excel para template '{parsed.name}'.",
        )
        AccessTemplateItem.objects.create(template=template, selection_set=selection_set, order=0)
        template.selection_set = selection_set
        template.save(update_fields=["selection_set"])

        if parsed.modules:
            SelectionSetModule.objects.bulk_create(
                [SelectionSetModule(selection_set=selection_set, module=module) for module in parsed.modules],
                batch_size=500,
            )
        if parsed.levels:
            SelectionSetLevel.objects.bulk_create(
                [SelectionSetLevel(selection_set=selection_set, level=level) for level in parsed.levels],
                batch_size=500,
            )
        if parsed.sublevels:
            SelectionSetSubLevel.objects.bulk_create(
                [SelectionSetSubLevel(selection_set=selection_set, sublevel=sublevel) for sublevel in parsed.sublevels],
                batch_size=500,
            )

        save_globals_for_selection_set(
            selection_set,
            action_items=parsed.action_items,
            matrix_items=[],
            payment_items=[],
        )

        results.append(
            SheetImportResult(
                sheet_name=parsed.name,
                template_id=template.pk,
                created=created,
                modules_selected=len(parsed.modules),
                sublevels_selected=len(parsed.sublevels),
                action_values_selected=parsed.action_values_selected,
            )
        )

    return WorkbookImportResult(results=results, warnings=warnings)
