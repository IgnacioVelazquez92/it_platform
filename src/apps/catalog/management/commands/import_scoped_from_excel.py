# src/apps/catalog/management/commands/import_scoped_from_excel.py
from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

from apps.catalog.models.permissions.scoped import (
    Company, Branch, Warehouse, CashRegister, ControlPanel, Seller
)


def _norm(s: object) -> str:
    return " ".join(str(s or "").strip().split())


def _h(s: object) -> str:
    return _norm(s).lower().replace(" ", "")


class Command(BaseCommand):
    help = "Importa Empresas, Sucursales, Depósitos, Cajas, Paneles y Vendedores desde Configuraciones.xlsx."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="Configuraciones.xlsx",
            help="Ruta al Excel (por defecto: Configuraciones.xlsx en la raíz del proyecto).",
        )
        parser.add_argument("--dry-run", action="store_true",
                            help="Simula sin escribir en DB.")

    @transaction.atomic
    def handle(self, *args, **options):
        dry_run: bool = options["dry_run"]

        # commands/ -> management -> catalog -> apps -> src -> it
        project_root = Path(__file__).resolve().parents[5]
        xlsx_path = Path(options["file"])
        if not xlsx_path.is_absolute():
            xlsx_path = project_root / xlsx_path

        if not xlsx_path.exists():
            raise CommandError(f"No existe el archivo: {xlsx_path}")

        wb = openpyxl.load_workbook(xlsx_path)

        created_company = 0
        created_branch = 0
        created_wh = 0
        created_cash = 0
        created_panel = 0
        created_seller = 0

        # -------------------------
        # 1) Empresas (hoja Empresa)
        # -------------------------
        if "Empresa" not in wb.sheetnames:
            raise CommandError("No existe la hoja 'Empresa'.")

        ws = wb["Empresa"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise CommandError("La hoja 'Empresa' está vacía.")

        hm = {_h(h): idx for idx, h in enumerate(header)}
        if "empresa" not in hm:
            raise CommandError(
                "La hoja 'Empresa' debe tener columna 'Empresa'.")

        for row in rows:
            name = _norm(row[hm["empresa"]] if hm["empresa"]
                         < len(row) else "")
            if not name:
                continue
            _, created = Company.objects.get_or_create(name=name)
            created_company += int(created)

        # --------------------------------
        # 2) Sucursales (hoja Sucursales)
        # --------------------------------
        if "Sucursales" not in wb.sheetnames:
            raise CommandError("No existe la hoja 'Sucursales'.")

        ws = wb["Sucursales"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise CommandError("La hoja 'Sucursales' está vacía.")

        hm = {_h(h): i for i, h in enumerate(header)}
        if "empresa" not in hm or "sucursal" not in hm:
            raise CommandError(
                "La hoja 'Sucursales' debe tener columnas: Empresa, Sucursal.")

        for row in rows:
            company_name = _norm(
                row[hm["empresa"]] if hm["empresa"] < len(row) else "")
            branch_name = _norm(row[hm["sucursal"]]
                                if hm["sucursal"] < len(row) else "")
            if not company_name or not branch_name:
                continue

            company, c_created = Company.objects.get_or_create(
                name=company_name)
            created_company += int(c_created)

            _, b_created = Branch.objects.get_or_create(
                company=company, name=branch_name)
            created_branch += int(b_created)

        # ------------------------------------------------
        # 3) Depósitos (hoja Depositos Hablitados)
        # ------------------------------------------------
        sheet_deps = "Depositos Habilitados"
        if sheet_deps not in wb.sheetnames:
            raise CommandError(f"No existe la hoja '{sheet_deps}'.")

        ws = wb[sheet_deps]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise CommandError(f"La hoja '{sheet_deps}' está vacía.")

        hm = {_h(h): i for i, h in enumerate(header)}
        if "empresa" not in hm or "sucursal" not in hm or "nombre" not in hm:
            raise CommandError(
                f"La hoja '{sheet_deps}' debe tener columnas: Empresa, Sucursal, Nombre.")

        for row in rows:
            company_name = _norm(
                row[hm["empresa"]] if hm["empresa"] < len(row) else "")
            branch_name = _norm(row[hm["sucursal"]]
                                if hm["sucursal"] < len(row) else "")
            wh_name = _norm(row[hm["nombre"]]
                            if hm["nombre"] < len(row) else "")

            if not company_name or not branch_name or not wh_name:
                continue

            company, c_created = Company.objects.get_or_create(
                name=company_name)
            created_company += int(c_created)

            branch, b_created = Branch.objects.get_or_create(
                company=company, name=branch_name)
            created_branch += int(b_created)

            _, w_created = Warehouse.objects.get_or_create(
                branch=branch, name=wh_name)
            created_wh += int(w_created)

        # -------------------------
        # 4) Cajas (hoja Cajas)
        # -------------------------
        sheet_cajas = "Cajas"
        if sheet_cajas not in wb.sheetnames:
            raise CommandError(f"No existe la hoja '{sheet_cajas}'.")

        ws = wb[sheet_cajas]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise CommandError(f"La hoja '{sheet_cajas}' está vacía.")

        hm = {_h(h): i for i, h in enumerate(header)}
        # columnas esperadas: Empresa, Nombre, Sucursal
        if "empresa" not in hm or "nombre" not in hm or "sucursal" not in hm:
            raise CommandError(
                f"La hoja '{sheet_cajas}' debe tener columnas: Empresa, Nombre, Sucursal.")

        for row in rows:
            company_name = _norm(
                row[hm["empresa"]] if hm["empresa"] < len(row) else "")
            cash_name = _norm(row[hm["nombre"]]
                              if hm["nombre"] < len(row) else "")
            branch_name = _norm(row[hm["sucursal"]]
                                if hm["sucursal"] < len(row) else "")

            if not company_name or not branch_name or not cash_name:
                continue

            company, c_created = Company.objects.get_or_create(
                name=company_name)
            created_company += int(c_created)

            branch, b_created = Branch.objects.get_or_create(
                company=company, name=branch_name)
            created_branch += int(b_created)

            _, cr_created = CashRegister.objects.get_or_create(
                branch=branch, name=cash_name)
            created_cash += int(cr_created)

        # ------------------------------
        # 5) Paneles (hoja Paneles Control)
        # ------------------------------
        sheet_panels = "Paneles Control"
        if sheet_panels not in wb.sheetnames:
            raise CommandError(f"No existe la hoja '{sheet_panels}'.")

        ws = wb[sheet_panels]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise CommandError(f"La hoja '{sheet_panels}' está vacía.")

        hm = {_h(h): i for i, h in enumerate(header)}
        # columnas esperadas: Empresa, Paneles
        if "empresa" not in hm or "paneles" not in hm:
            raise CommandError(
                f"La hoja '{sheet_panels}' debe tener columnas: Empresa, Paneles.")

        for row in rows:
            company_name = _norm(
                row[hm["empresa"]] if hm["empresa"] < len(row) else "")
            panel_name = _norm(row[hm["paneles"]]
                               if hm["paneles"] < len(row) else "")
            if not company_name or not panel_name:
                continue

            company, c_created = Company.objects.get_or_create(
                name=company_name)
            created_company += int(c_created)

            _, p_created = ControlPanel.objects.get_or_create(
                company=company, name=panel_name)
            created_panel += int(p_created)

        # -------------------------
        # 6) Vendedores (hoja Vendedores)
        # -------------------------
        sheet_sellers = "Vendedores"
        if sheet_sellers not in wb.sheetnames:
            raise CommandError(f"No existe la hoja '{sheet_sellers}'.")

        ws = wb[sheet_sellers]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise CommandError(f"La hoja '{sheet_sellers}' está vacía.")

        hm = {_h(h): i for i, h in enumerate(header)}
        # columnas esperadas: Empresa, Vendedor
        if "empresa" not in hm or "vendedor" not in hm:
            raise CommandError(
                f"La hoja '{sheet_sellers}' debe tener columnas: Empresa, Vendedor.")

        for row in rows:
            company_name = _norm(
                row[hm["empresa"]] if hm["empresa"] < len(row) else "")
            seller_name = _norm(row[hm["vendedor"]]
                                if hm["vendedor"] < len(row) else "")
            if not company_name or not seller_name:
                continue

            company, c_created = Company.objects.get_or_create(
                name=company_name)
            created_company += int(c_created)

            _, s_created = Seller.objects.get_or_create(
                company=company, name=seller_name)
            created_seller += int(s_created)

        if dry_run:
            transaction.set_rollback(True)
            self.stdout.write(self.style.WARNING(
                "DRY-RUN: no se guardó nada en la DB."))

        self.stdout.write(
            self.style.SUCCESS(
                "OK. "
                f"Empresas: {created_company} | "
                f"Sucursales: {created_branch} | "
                f"Depósitos: {created_wh} | "
                f"Cajas: {created_cash} | "
                f"Paneles: {created_panel} | "
                f"Vendedores: {created_seller}"
            )
        )
