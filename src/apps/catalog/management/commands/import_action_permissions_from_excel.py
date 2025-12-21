from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

from apps.catalog.models.permissions.global_ops import (
    ActionPermission,
    ActionValueType,
    MatrixPermission,
    PaymentMethodPermission,
)


def _norm(s: object) -> str:
    return " ".join(str(s or "").strip().split())


def _h(s: object) -> str:
    return _norm(s).lower().replace(" ", "")


TYPE_MAP = {
    "bool": ActionValueType.BOOL,
    "booleano": ActionValueType.BOOL,

    "entero": ActionValueType.INT,
    "int": ActionValueType.INT,

    "decimal": ActionValueType.DECIMAL,
    "número": ActionValueType.DECIMAL,
    "numero": ActionValueType.DECIMAL,

    "porcentaje": ActionValueType.PERCENT,
    "%": ActionValueType.PERCENT,

    "texto": ActionValueType.TEXT,
    "text": ActionValueType.TEXT,
}


class Command(BaseCommand):
    help = "Importa permisos globales desde Configuraciones.xlsx (acciones, matriz y medios de pago)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="Configuraciones.xlsx",
            help="Ruta al Excel (por defecto: Configuraciones.xlsx en la raíz del proyecto).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula sin escribir en DB.",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        dry_run: bool = options["dry_run"]

        project_root = Path(__file__).resolve().parents[5]
        xlsx_path = Path(options["file"])
        if not xlsx_path.is_absolute():
            xlsx_path = project_root / xlsx_path

        if not xlsx_path.exists():
            raise CommandError(f"No existe el archivo: {xlsx_path}")

        wb = openpyxl.load_workbook(xlsx_path)

        # ---------------------------------------------------------
        # 1) Permisos de Acciones
        # ---------------------------------------------------------
        sheet = "Permisos de Acciones"
        if sheet not in wb.sheetnames:
            raise CommandError(f"No existe la hoja '{sheet}'.")

        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)

        hm = {_h(h): i for i, h in enumerate(header or [])}
        for col in ("tipo", "acciones", "permiso"):
            if col not in hm:
                raise CommandError(f"Falta columna '{col}' en hoja '{sheet}'.")

        created = updated = skipped = processed = 0

        for row in rows:
            processed += 1

            group = _norm(row[hm["tipo"]])
            action = _norm(row[hm["acciones"]])
            vt_raw = _norm(row[hm["permiso"]])

            if not group or not action or not vt_raw:
                skipped += 1
                continue

            value_type = TYPE_MAP.get(vt_raw.lower())
            if not value_type:
                raise CommandError(f"Tipo de permiso desconocido: '{vt_raw}'")

            obj, was_created = ActionPermission.objects.get_or_create(
                group=group,
                action=action,
                defaults={"value_type": value_type},
            )
            if was_created:
                created += 1
            elif obj.value_type != value_type:
                obj.value_type = value_type
                obj.save(update_fields=["value_type"])
                updated += 1

        # ---------------------------------------------------------
        # 2) Permisos (matriz simplificada)
        # ---------------------------------------------------------
        sheet_perm = "Permisos"
        created_matrix = updated_matrix = skipped_matrix = 0

        if sheet_perm in wb.sheetnames:
            ws = wb[sheet_perm]
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)

            hm = {_h(h): i for i, h in enumerate(header or [])}
            if "permisos" not in hm:
                raise CommandError(
                    "La hoja 'Permisos' debe tener columna 'Permisos'.")

            COLS = {
                "crear": "can_create",
                "modificar": "can_update",
                "autorizar": "can_authorize",
                "cerrar": "can_close",
                "anular": "can_cancel",
                "actualizavigencia": "can_update_validity",
            }

            for row in rows:
                name = _norm(row[hm["permisos"]])
                if not name:
                    skipped_matrix += 1
                    continue

                values = {}
                for col_key, field in COLS.items():
                    if col_key in hm:
                        cell = _norm(row[hm[col_key]])
                        values[field] = bool(cell)

                perm, created_p = MatrixPermission.objects.get_or_create(
                    name=name, defaults=values
                )

                if created_p:
                    created_matrix += 1
                else:
                    for k, v in values.items():
                        setattr(perm, k, v)
                    perm.save()
                    updated_matrix += 1

        # ---------------------------------------------------------
        # 3) Medios de Pago
        # ---------------------------------------------------------
        sheet_mp = "Medios de Pago"
        created_mp = skipped_mp = 0

        if sheet_mp in wb.sheetnames:
            ws = wb[sheet_mp]
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)

            hm = {_h(h): i for i, h in enumerate(header or [])}
            if "mediosdepago" not in hm:
                raise CommandError(
                    "La hoja 'Medios de Pago' debe tener columna 'Medios de Pago'.")

            for row in rows:
                name = _norm(row[hm["mediosdepago"]])
                if not name:
                    skipped_mp += 1
                    continue
                _, c = PaymentMethodPermission.objects.get_or_create(name=name)
                created_mp += int(c)

        if dry_run:
            transaction.set_rollback(True)
            self.stdout.write(self.style.WARNING(
                "DRY-RUN: no se guardó nada en la DB."))

        self.stdout.write(
            self.style.SUCCESS(
                "OK. "
                f"Acciones -> Procesadas: {processed} | Saltadas: {skipped} | "
                f"Creadas: {created} | Actualizadas: {updated} || "
                f"Matriz -> Saltadas: {skipped_matrix} | "
                f"Creadas: {created_matrix} | Actualizadas: {updated_matrix} || "
                f"Medios de Pago -> Saltadas: {skipped_mp} | Creadas: {created_mp}"
            )
        )
